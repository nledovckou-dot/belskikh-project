import openpyxl
import re
import copy

# ============================================================
# 1. Load workbooks
# ============================================================
src_wb = openpyxl.load_workbook('/Users/n-a-ledovskoy/Downloads/Бельских_проект/Выгрузка результатов_ИИ.xlsx')
tgt_wb = openpyxl.load_workbook('/Users/n-a-ledovskoy/Downloads/Бельских_проект/Сводная форма отчета_ИИ.xlsx')

src_ws = src_wb['с трейд-ин']
tgt_ws = tgt_wb['АСП Продажи_с трейд-ин']

NUM_QUESTIONNAIRES = 4  # rows 2-5 in source
SCORE_START_COL = 5    # E = col 5 (first questionnaire scores)
COMMENT_START_COL = 16  # P = col 16 (first questionnaire comments)

# ============================================================
# 2. Read source data
# ============================================================
# Headers from row 1
src_headers = {}
for c in range(1, src_ws.max_column + 1):
    v = src_ws.cell(1, c).value
    if v is not None:
        src_headers[c] = str(v).strip()

# Data from rows 2-5 (4 questionnaires)
src_data = []
for r in range(2, 2 + NUM_QUESTIONNAIRES):
    row_data = {}
    for c in range(1, src_ws.max_column + 1):
        v = src_ws.cell(r, c).value
        row_data[c] = v
    src_data.append(row_data)

print(f"Source: {len(src_headers)} header columns, {len(src_data)} questionnaires")

# ============================================================
# 3. Build source question mapping
# ============================================================
# Identify question columns and their comment columns
# Pattern: question col followed by "Комментарий" col
# Section headers have % values, sub-section headers have None values

def normalize_text(t):
    """Normalize text for matching"""
    if t is None:
        return ""
    t = str(t).strip()
    # Remove extra spaces
    t = re.sub(r'\s+', ' ', t)
    # Remove trailing punctuation
    t = t.rstrip('.')
    # Lowercase
    t = t.lower()
    # Normalize quotes
    t = t.replace('"', "'").replace('"', "'").replace('"', "'").replace('«', "'").replace('»', "'")
    return t

# Find question columns: cols after 37 that have headers not starting with "Комментарий"
# and are not section/sub-section headers
question_cols = []  # list of (question_col, comment_col_or_None, header_text)

# Section headers in source (cols with % values in data)
section_cols = set()
subsection_cols = set()

for c in sorted(src_headers.keys()):
    if c <= 37:
        continue
    header = src_headers[c]
    
    # Check if this is a "Комментарий" column
    if header.startswith('Комментарий'):
        continue
    
    # Check data values to determine type
    values = [src_data[i].get(c) for i in range(NUM_QUESTIONNAIRES)]
    non_none_values = [v for v in values if v is not None]
    
    if not non_none_values:
        # All None = sub-section header
        subsection_cols.add(c)
        continue
    
    # Check if values are percentages (section summary)
    has_pct = any(isinstance(v, str) and '%' in str(v) for v in non_none_values)
    if has_pct:
        section_cols.add(c)
        continue
    
    # This is a question column - find its comment column
    comment_col = None
    next_c = c + 1
    if next_c in src_headers and src_headers[next_c].startswith('Комментарий'):
        # Make sure it's not "Комментарий к разделу"
        if not src_headers[next_c].startswith('Комментарий к разделу'):
            comment_col = next_c
    
    question_cols.append((c, comment_col, header))

print(f"\nFound {len(question_cols)} questions in source")
print(f"Section summary cols: {len(section_cols)}, Sub-section cols: {len(subsection_cols)}")

# ============================================================
# 4. Build target row mapping
# ============================================================
# Read all question texts from target col 4 (rows 52-148)
target_questions = {}  # normalized_text -> row_number
target_questions_raw = {}  # row -> raw text

for r in range(52, tgt_ws.max_row + 1):
    text = tgt_ws.cell(r, 4).value
    if text is not None:
        text_str = str(text).strip()
        norm = normalize_text(text_str)
        target_questions[norm] = r
        target_questions_raw[r] = text_str

print(f"\nFound {len(target_questions)} question rows in target")

# ============================================================
# 5. Match source questions to target rows
# ============================================================
matched = []
unmatched_src = []

for q_col, c_col, header in question_cols:
    norm_header = normalize_text(header)
    
    if norm_header in target_questions:
        target_row = target_questions[norm_header]
        matched.append((q_col, c_col, header, target_row))
    else:
        # Try partial matching - find best match
        best_match = None
        best_score = 0
        for norm_tgt, tgt_row in target_questions.items():
            # Simple overlap score
            words_src = set(norm_header.split())
            words_tgt = set(norm_tgt.split())
            if not words_src or not words_tgt:
                continue
            overlap = len(words_src & words_tgt) / max(len(words_src), len(words_tgt))
            if overlap > best_score:
                best_score = overlap
                best_match = (norm_tgt, tgt_row)
        
        if best_match and best_score >= 0.7:
            matched.append((q_col, c_col, header, best_match[1]))
        else:
            unmatched_src.append((q_col, c_col, header, best_match, best_score))

print(f"\nMatched: {len(matched)}")
if unmatched_src:
    print(f"Unmatched: {len(unmatched_src)}")
    for q_col, c_col, header, best, score in unmatched_src:
        print(f"  col {q_col}: '{header[:60]}...' best={best}, score={score:.2f}")

# ============================================================
# 6. Fill meta info (rows 3-39)
# ============================================================
# Source cols 1-37 map to target rows 3-39
# The mapping is by label matching
meta_mapping = {}  # target_row -> source_col

# Build target meta labels
target_meta = {}
for r in range(3, 40):
    label = tgt_ws.cell(r, 4).value
    if label:
        target_meta[normalize_text(str(label))] = r

# Build source meta labels  
source_meta = {}
for c in range(1, 38):
    if c in src_headers:
        source_meta[normalize_text(src_headers[c])] = c

# Match
for norm_label, tgt_row in target_meta.items():
    if norm_label in source_meta:
        meta_mapping[tgt_row] = source_meta[norm_label]

print(f"\nMeta fields matched: {len(meta_mapping)}")

# Fill meta data
for tgt_row, src_col in meta_mapping.items():
    for q_idx in range(NUM_QUESTIONNAIRES):
        value = src_data[q_idx].get(src_col)
        col = SCORE_START_COL + q_idx  # E, F, G, H
        tgt_ws.cell(tgt_row, col, value)

print("Meta data filled")

# ============================================================
# 7. Fill question scores and comments
# ============================================================
for q_col, c_col, header, target_row in matched:
    for q_idx in range(NUM_QUESTIONNAIRES):
        # Score
        score = src_data[q_idx].get(q_col)
        score_col = SCORE_START_COL + q_idx  # E, F, G, H (5,6,7,8)
        tgt_ws.cell(target_row, score_col, score)
        
        # Comment
        if c_col is not None:
            comment = src_data[q_idx].get(c_col)
            comment_col = COMMENT_START_COL + q_idx  # P, Q, R, S (16,17,18,19)
            tgt_ws.cell(target_row, comment_col, comment)

print(f"Question data filled: {len(matched)} questions x {NUM_QUESTIONNAIRES} questionnaires")

# ============================================================
# 8. Fill section-level comments (Комментарий к разделу)
# ============================================================
# These go into the section header rows in the target
# Source has "Комментарий к разделу Звонок" etc.
section_comment_mapping = {
    'Комментарий к разделу Звонок': 51,  # Звонок section header
    'Комментарий к разделу Парковка и инфраструктура': 67,
    'Комментарий к разделу Презентация в автосалоне': 77,
    'Комментарий к разделу Трейд-ин и КСО': 102,
    'Комментарий к разделу Коммерческое предложение': 135,
    'Комментарий к разделу Завершающие шаги': 142,
}

for c in sorted(src_headers.keys()):
    header = src_headers[c]
    if header.startswith('Комментарий к разделу'):
        if header in section_comment_mapping:
            tgt_row = section_comment_mapping[header]
            for q_idx in range(NUM_QUESTIONNAIRES):
                value = src_data[q_idx].get(c)
                if value is not None:
                    comment_col = COMMENT_START_COL + q_idx
                    tgt_ws.cell(tgt_row, comment_col, value)
            print(f"  Section comment '{header}' → row {tgt_row}")

# ============================================================
# 9. Clear old sample data at col 26 (rows 3-39)
# ============================================================
for r in range(3, 40):
    old_val = tgt_ws.cell(r, 26).value
    if old_val is not None:
        tgt_ws.cell(r, 26, None)
        print(f"  Cleared old data at ({r}, 26): {repr(old_val)[:50]}")

# ============================================================
# 10. Save
# ============================================================
output_path = '/Users/n-a-ledovskoy/Downloads/Бельских_проект/Сводная форма отчета_ИИ.xlsx'
tgt_wb.save(output_path)
print(f"\nSaved to: {output_path}")
print("DONE!")
